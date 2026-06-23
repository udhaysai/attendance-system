from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from .models import Student
from .forms import StudentForm
from datetime import date
from .models import Student, Attendance
from datetime import timedelta
from django.http import HttpResponse
import openpyxl
from datetime import date, timedelta
from django.contrib.sessions.models import Session
from .models import UserSession





def login_view(request):

    if request.method == 'POST':

        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user is not None:

            old_session = UserSession.objects.filter(user=user).first()

            if old_session and old_session.session_key:
                Session.objects.filter(
                    session_key=old_session.session_key
                ).delete()

            login(request, user)

            UserSession.objects.update_or_create(
                user=user,
                defaults={
                    'session_key': request.session.session_key
                }
            )

            return redirect('dashboard')

        return render(
            request,
            'login.html',
            {'error': 'Invalid Username or Password'}
        )

    return render(request, 'login.html')


@login_required
def dashboard(request):

    total_students = Student.objects.count()

    today = date.today()

    present_count = Attendance.objects.filter(
        date=today,
        status='Present'
    ).count()

    absent_count = Attendance.objects.filter(
        date=today,
        status='Absent'
    ).count()

    attendance_percentage = 0

    if total_students > 0:
        attendance_percentage = (
            present_count / total_students
        ) * 100

    context = {
        'total_students': total_students,
        'present_count': present_count,
        'absent_count': absent_count,
        'attendance_percentage': round(attendance_percentage, 2)
    }

    return render(
        request,
        'dashboard.html',
        context
    )

@login_required
def logout_view(request):

    UserSession.objects.filter(user=request.user).delete()

    logout(request)

    return redirect('login')





@login_required
def student_list(request):

    query = request.GET.get('q')

    if query:

        students = Student.objects.filter(
            name__icontains=query
        ) | Student.objects.filter(
            roll_no__icontains=query
        )

    else:
        students = Student.objects.all()

    return render(
        request,
        'students/list.html',
        {
            'students': students
        }
    )


@login_required
def student_add(request):

    if request.method == 'POST':

        roll_no = request.POST.get('roll_no', '').strip()
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        department = request.POST.get('department', '').strip()

        if not roll_no or not name or not email or not phone or not department:
            return render(
                request,
                'students/add.html',
                {
                    'error': 'All fields are required.',
                    'roll_no': roll_no,
                    'name': name,
                    'email': email,
                    'phone': phone,
                    'department': department
                }
            )

        Student.objects.create(
            roll_no=roll_no,
            name=name,
            email=email,
            phone=phone,
            department=department
        )

        return redirect('student_list')

    return render(request, 'students/add.html')



@login_required
def student_edit(request, id):
    student = Student.objects.get(id=id)

    if request.method == 'POST':
        student.roll_no = request.POST['roll_no']
        student.name = request.POST['name']
        student.email = request.POST['email']
        student.phone = request.POST['phone']
        student.department = request.POST['department']
        student.save()

        return redirect('student_list')

    return render(request, 'students/edit.html', {'student': student})


@login_required
def student_delete(request, id):
    student = Student.objects.get(id=id)
    student.delete()

    return redirect('student_list')



@login_required
def mark_attendance(request):

    students = Student.objects.all()
    today = date.today()

    if request.method == 'POST':

        for student in students:

            status = request.POST.get(
                f'status_{student.id}'
            )

            if status:

                Attendance.objects.update_or_create(
                    student=student,
                    date=today,
                    defaults={
                        'status': status
                    }
                )

        return redirect('dashboard')

    return render(
        request,
        'attendance/mark.html',
        {
            'students': students,
            'today': today
        }
    )





@login_required
def attendance_history(request):

    selected_date = request.GET.get('date')

    records = None

    if selected_date:
        records = Attendance.objects.filter(
            date=selected_date
        )

    return render(
        request,
        'attendance/history.html',
        {
            'records': records
        }
    )



@login_required
def daily_report(request):

    records = None

    selected_date = request.GET.get('date')

    if selected_date:

        records = Attendance.objects.filter(
            date=selected_date
        )

    return render(
        request,
        'reports/daily.html',
        {
            'records': records
        }
    )

@login_required
def weekly_report(request):

    records = None

    start_date = request.GET.get('start_date')

    end_date = None

    if start_date:

        start = date.fromisoformat(start_date)

        end = start + timedelta(days=6)

        end_date = end

        records = Attendance.objects.filter(
            date__range=[start, end]
        )

    return render(
        request,
        'reports/weekly.html',
        {
            'records': records,
            'end_date': end_date
        }
    )


@login_required
def monthly_report(request):

    records = None

    month = request.GET.get('month')

    if month:

        year, month_num = month.split('-')

        records = Attendance.objects.filter(
            date__year=year,
            date__month=month_num
        )

    return render(
        request,
        'reports/monthly.html',
        {
            'records': records
        }
    )

@login_required
def export_daily_excel(request):

    selected_date = request.GET.get('date')

    records = Attendance.objects.filter(
        date=selected_date
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Daily Report"

    sheet.append([
        "Roll No",
        "Name",
        "Date",
        "Status"
    ])

    for record in records:
        sheet.append([
            record.student.roll_no,
            record.student.name,
            str(record.date),
            record.status
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=daily_attendance_report.xlsx'

    workbook.save(response)

    return response






@login_required
def export_weekly_excel(request):

    start_date = request.GET.get('start_date')

    start = date.fromisoformat(start_date)
    end = start + timedelta(days=6)

    records = Attendance.objects.filter(
        date__range=[start, end]
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Weekly Report"

    sheet.append(["Roll No", "Name", "Date", "Status"])

    for record in records:
        sheet.append([
            record.student.roll_no,
            record.student.name,
            str(record.date),
            record.status
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=weekly_attendance_report.xlsx'

    workbook.save(response)

    return response




@login_required
def export_monthly_excel(request):

    month = request.GET.get('month')

    year, month_num = month.split('-')

    records = Attendance.objects.filter(
        date__year=year,
        date__month=month_num
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Monthly Report"

    sheet.append(["Roll No", "Name", "Date", "Status"])

    for record in records:
        sheet.append([
            record.student.roll_no,
            record.student.name,
            str(record.date),
            record.status
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=monthly_attendance_report.xlsx'

    workbook.save(response)

    return response



